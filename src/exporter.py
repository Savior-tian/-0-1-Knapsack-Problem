"""
导出模块：将 D{0-1}KP 求解结果导出为 TXT 或 Excel 文件。

TXT 格式：
    包含实例名称、容量、最优值、求解时间以及所选物品列表。

Excel 格式：
    Sheet1 "摘要"：基本信息。
    Sheet2 "所选物品"：每行一件选中物品的详情。
"""

import csv
import datetime
from typing import List

from src.data_parser import DKPInstance
from src.dp_solver import SolveResult, solve
from src.sorter import get_sorted_ratios


def _create_header_style(font_cls, fill_cls):
    """创建 Excel 统一表头样式。"""
    header_font = font_cls(bold=True, color="FFFFFF")
    header_fill = fill_cls(fill_type="solid", fgColor="457B9D")
    return header_font, header_fill


def _fill_summary_sheet(
    ws_summary, instance: DKPInstance, result: SolveResult, header_font, header_fill
) -> None:
    ws_summary.title = "摘要"

    summary_data = [
        ("项目", "内容"),
        ("实例名称", result.instance_name),
        ("组数 (N)", instance.num_groups),
        ("背包容量", instance.capacity),
        ("最优总价值", result.optimal_value),
        ("求解耗时 (ms)", f"{result.solve_time_ms:.4f}"),
        ("选中物品数", len(result.selected_items)),
        ("导出时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for row_idx, (key, val) in enumerate(summary_data, start=1):
        cell_k = ws_summary.cell(row=row_idx, column=1, value=key)
        cell_v = ws_summary.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            cell_k.font = header_font
            cell_k.fill = header_fill
            cell_v.font = header_font
            cell_v.fill = header_fill

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30


def _fill_items_sheet(
    ws_items, result: SolveResult, header_font, header_fill, alignment_cls, font_cls
) -> None:
    item_headers = ["组编号", "组内编号", "价值", "重量"]
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_cls(horizontal="center")

    total_weight = 0
    for row_idx, (group_idx, item_in_group, profit, weight) in enumerate(
        result.selected_items, start=2
    ):
        ws_items.cell(row=row_idx, column=1, value=group_idx + 1)
        ws_items.cell(row=row_idx, column=2, value=item_in_group + 1)
        ws_items.cell(row=row_idx, column=3, value=profit)
        ws_items.cell(row=row_idx, column=4, value=weight)
        total_weight += weight

    last_row = len(result.selected_items) + 2
    ws_items.cell(row=last_row, column=1, value="合计").font = font_cls(bold=True)
    ws_items.cell(row=last_row, column=3, value=result.optimal_value).font = font_cls(
        bold=True
    )
    ws_items.cell(row=last_row, column=4, value=total_weight).font = font_cls(bold=True)

    for col in ["A", "B", "C", "D"]:
        ws_items.column_dimensions[col].width = 14


def _fill_sort_sheet(
    ws_sort,
    instance: DKPInstance,
    sorted_order: List[int],
    header_font,
    header_fill,
    alignment_cls,
) -> None:
    sort_headers = ["排序名次", "原始组编号", "第3件价值", "第3件重量", "价值/重量比"]
    for col_idx, header in enumerate(sort_headers, start=1):
        cell = ws_sort.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_cls(horizontal="center")

    ratios = get_sorted_ratios(instance, sorted_order)
    for rank, (orig_idx, ratio) in enumerate(zip(sorted_order, ratios), start=1):
        grp = instance.groups[orig_idx]
        ws_sort.cell(row=rank + 1, column=1, value=rank)
        ws_sort.cell(row=rank + 1, column=2, value=orig_idx + 1)
        ws_sort.cell(row=rank + 1, column=3, value=grp.profits[2])
        ws_sort.cell(row=rank + 1, column=4, value=grp.weights[2])
        ws_sort.cell(row=rank + 1, column=5, value=round(ratio, 6))

    for col in ["A", "B", "C", "D", "E"]:
        ws_sort.column_dimensions[col].width = 16


def export_txt(
    instance: DKPInstance,
    result: SolveResult,
    out_path: str,
) -> None:
    """
    将求解结果导出为纯文本文件。

    Args:
        instance: 原始实例（用于输出物品详情）。
        result:   SolveResult 求解结果。
        out_path: 输出文件路径。
    """
    lines = [
        "=" * 60,
        "  D{0-1} 背包问题求解结果",
        "=" * 60,
        f"实例名称  : {result.instance_name}",
        f"组数 (N)  : {instance.num_groups}",
        f"背包容量  : {instance.capacity}",
        f"最优总价值: {result.optimal_value}",
        f"求解耗时  : {result.solve_time_ms:.4f} ms",
        f"导出时间  : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "-" * 60,
        "所选物品列表（共 {} 件）:".format(len(result.selected_items)),
        "-" * 60,
        f"{'组编号':>6}  {'组内编号':>8}  {'价值':>8}  {'重量':>8}",
    ]

    total_weight = 0
    for group_idx, item_in_group, profit, weight in result.selected_items:
        lines.append(
            f"{group_idx + 1:>6}  {item_in_group + 1:>8}  {profit:>8}  {weight:>8}"
        )
        total_weight += weight

    lines += [
        "-" * 60,
        f"{'合计':>6}  {'':>8}  {result.optimal_value:>8}  {total_weight:>8}",
        "=" * 60,
    ]

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def export_excel(
    instance: DKPInstance,
    result: SolveResult,
    sorted_order: List[int],
    out_path: str,
) -> None:
    """
    将求解结果导出为 Excel 文件（.xlsx）。
    需要 openpyxl 库。

    Args:
        instance:     原始实例。
        result:       SolveResult 求解结果。
        sorted_order: 排序后各组的原始索引列表（来自 sort_groups_by_ratio）。
        out_path:     输出文件路径（.xlsx）。
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            "导出 Excel 需要安装 openpyxl，请运行：pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_items = wb.create_sheet("所选物品")
    ws_sort = wb.create_sheet("按比值排序")

    header_font, header_fill = _create_header_style(Font, PatternFill)
    _fill_summary_sheet(ws_summary, instance, result, header_font, header_fill)
    _fill_items_sheet(ws_items, result, header_font, header_fill, Alignment, Font)
    _fill_sort_sheet(
        ws_sort,
        instance,
        sorted_order,
        header_font,
        header_fill,
        Alignment,
    )

    wb.save(out_path)


def export_batch_summary_csv(instances: list[DKPInstance], out_path: str) -> None:
    """批量求解实例并导出汇总 CSV。"""
    headers = [
        "instance_name",
        "num_groups",
        "capacity",
        "optimal_value",
        "solve_time_ms",
        "selected_items_count",
        "total_weight",
        "weight_utilization",
    ]

    with open(out_path, "w", encoding="utf-8-sig", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)

        for instance in instances:
            result = solve(instance)
            total_weight = sum(item[3] for item in result.selected_items)
            if instance.capacity > 0:
                utilization = total_weight / instance.capacity
            else:
                utilization = 0.0
            writer.writerow(
                [
                    instance.name,
                    instance.num_groups,
                    instance.capacity,
                    result.optimal_value,
                    round(result.solve_time_ms, 4),
                    len(result.selected_items),
                    total_weight,
                    round(utilization, 6),
                ]
            )
